import os
from get_commands import get_show_version
from get_commands import get_show_int_status
from get_commands import get_show_cdp_nei_det
import openpyxl


def show_version_to_excel(folder):

    # Change to Log File Directory
    os.chdir(folder)

    # Create site xlsx if not already created
    if not os.path.exists('site-name.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('site-name.xlsx')
        wb.close

    # Open workbook and delete sh_ver sheet if it exists
    wb = openpyxl.load_workbook(filename='site-name.xlsx')
    if 'sh_ver' in wb.get_sheet_names():
        ws = wb.get_sheet_by_name('sh_ver')
        wb.remove_sheet(ws)

    # Create sh_ver sheet and make active
    wb.create_sheet(title='sh_ver')
    sh_ver_sheet = wb.get_sheet_by_name('sh_ver')

    # For every file in Log Directory ending with .log
    for fn in os.listdir(os.getcwd()):

        # Get show ver header and show ver data
        if fn.endswith(".log"):
            sh_ver_header, sh_ver_data = get_show_version(fn)

            # If row sheet length is 1, write header else write data
            if sh_ver_sheet.max_row == 1:
                sh_ver_sheet.append(sh_ver_header)

            # For each line in list write data excel row
            for line in sh_ver_data:
                sh_ver_sheet.append(line)

    # Save workbook
    wb.save('site-name.xlsx')


def show_int_stat_to_excel(folder):

    # Change to Log File Directory
    os.chdir(folder)

    # Create site xlsx if not already created
    if not os.path.exists('site-name.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('site-name.xlsx')
        wb.close

    # Open workbook and delete sh_int_stat sheet if it exists
    wb = openpyxl.load_workbook(filename='site-name.xlsx')
    if 'sh_int_stat' in wb.get_sheet_names():
        ws = wb.get_sheet_by_name('sh_int_stat')
        wb.remove_sheet(ws)

    # Create sh_int_stat sheet and make active
    wb.create_sheet(title='sh_int_stat')
    sh_int_stat_sheet = wb.get_sheet_by_name('sh_int_stat')

    # For every file in Log Directory ending with .log
    for fn in os.listdir(os.getcwd()):

        # Get sh_ver header and data, sh_int_stat header and data
        if fn.endswith(".log"):
            sh_int_stat_header, sh_int_stat_data = get_show_int_status(fn)
            sh_ver_header, sh_ver_data = get_show_version(fn)

            # If row sheet length is 1, insert HOSTNAME to header and write header
            if sh_int_stat_sheet.max_row == 1:
                sh_int_stat_header.insert(0, sh_ver_header[2])
                sh_int_stat_sheet.append(sh_int_stat_header)

            # For each line in list insert <hostname> and write data to excel row
            for line in sh_int_stat_data:
                line.insert(0, sh_ver_data[0][2])
                sh_int_stat_sheet.append(line)

    # Save workbook
    wb.save('site-name.xlsx')

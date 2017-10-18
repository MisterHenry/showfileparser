import sys
import os
import re
import textfsm
import openpyxl
import datetime

def log_file_to_show_ver(logfile):

    # Load the input file to a variable
    input_file = open(logfile)
    raw_text_data = input_file.read()
    input_file.close()

    # Get SH VERSION Raw Data
    raw_version_data = re.search(r'#sh ver.*#sh mod', raw_text_data, re.DOTALL)

    # Run SHOW VERSION through the FSM.
    show_version_template = open("../textfsm_templates/cisco_ios_show_version.template")
    show_version_table = textfsm.TextFSM(show_version_template)
    show_version_parsed = show_version_table.ParseText(raw_version_data.group(0))

    return show_version_table.header, show_version_parsed


def log_file_to_show_int_stat(logfile):

    # Load the input file to a variable
    input_file = open(logfile)
    raw_text_data = input_file.read()
    input_file.close()

    # Get SH INT STAT Raw Data
    raw_interfaces_status_data = re.search(r'#sh int status.*#sh int trunk', raw_text_data, re.DOTALL)

    # Run SHOW INTERFACES STATUS through the FSM.
    show_interfaces_status_template = open("../textfsm_templates/cisco_ios_show_interfaces_status.template")
    show_interfaces_status_table = textfsm.TextFSM(show_interfaces_status_template)
    show_interfaces_status_parsed = show_interfaces_status_table.ParseText(raw_interfaces_status_data.group(0))

    return show_interfaces_status_table.header, show_interfaces_status_parsed


def log_file_to_show_cdp_nei_det(logfile):

    # Load the input file to a variable
    input_file = open(logfile)
    raw_text_data = input_file.read()
    input_file.close()

    # Get SH CDP NEI Raw Data
    raw_cdp_nei_det_data = re.search(r'#sh cdp nei det.*#sh int desc', raw_text_data, re.DOTALL)

    # Run SHOW CDP NEI DET through the FSM.
    show_cdp_nei_det_template = open("../textfsm_templates/cisco_ios_show_cdp_neighbors_detail.template")
    show_cdp_nei_det_table = textfsm.TextFSM(show_cdp_nei_det_template)
    show_cdp_nei_det_parsed = show_cdp_nei_det_table.ParseText(raw_cdp_nei_det_data.group(0))

    # Normalize Interface Name

    for line in show_cdp_nei_det_parsed:
        for i in range(0, len(line)):
            if 'GigabitEthernet' in line[i]:
                line[i] = re.sub('^GigabitEthernet', 'Gi', line[i])

            if 'TenGigabitEthernet' in line[i]:
                line[i] = re.sub('^TenGigabitEthernet', 'Te', line[i])

    return show_cdp_nei_det_table.header, show_cdp_nei_det_parsed


def log_file_to_matrix(logfile):

    sh_ver_header = []
    sh_int_stat_header = []
    sh_cdp_nei_det_header = []
    sh_int_stat_data = []
    matrix_data = []

    # Get show header and show  data
    if logfile.endswith(".log"):
        sh_ver_header, sh_ver_data = log_file_to_show_ver(logfile)
        sh_int_stat_header, sh_int_stat_data = log_file_to_show_int_stat(logfile)
        sh_cdp_nei_det_header, sh_cdp_nei_det_data = log_file_to_show_cdp_nei_det(logfile)

    # Define header --> list of 21 items
    matrix_header = sh_ver_header + sh_int_stat_header + sh_cdp_nei_det_header

    # Add cdp info for matching interface lines
    for line in sh_int_stat_data:
        for entry in sh_cdp_nei_det_data:
            if line[0] == entry[4]:
                line.extend(entry)
    # Add sh_ver_data to each line in list
    for line in sh_int_stat_data:
        matrix_data.append(sh_ver_data[0] + line)

    # Matrix header --> list -- len 21
    # Matrix data --> list of lists -- len 15 or 21
    return matrix_header, matrix_data


def log_folder_to_matrix(logfolder):

    # Change to Log File Directory
    os.chdir(logfolder)

    # List to hold all matrix data
    # ['VERSION', 'ROMMON', 'HOSTNAME', 'UPTIME', 'RUNNING_IMAGE', 'HARDWARE', 'SERIAL', 'CONFIG_REGISTER',
    # 'PORT', 'NAME', 'STATUS', 'VLAN', 'DUPLEX', 'SPEED', 'TYPE',
    # 'DESTINATION_HOST', 'MANAGEMENT_IP', 'PLATFORM', 'REMOTE_PORT', 'LOCAL_PORT', 'SOFTWARE_VERSION']
    log_folder_matrix = []

    # For each log file  get matrix header and matrix data
    for logfile in os.listdir(os.getcwd()):
        if logfile.endswith(".log"):
            header, data = log_file_to_matrix(logfile)
            # If matrix_all is empty add header

            if len(log_folder_matrix) == 0:
                log_folder_matrix.append(header)
            for line in data:
                log_folder_matrix.append(line)

    # Return list of lists. One per pyhsical interface per device
    return log_folder_matrix


def show_version_to_excel(site, logfolder):

    # Change to Log File Directory
    os.chdir(logfolder)

    # Create site xlsx if not already created
    if not os.path.exists('{}-show-info.xlsx'.format(site)):
        wb = openpyxl.Workbook()
        wb.save('{}-show-info.xlsx'.format(site))
        wb.close

    # Open workbook and delete sh_ver sheet if it exists
    wb = openpyxl.load_workbook(filename='{}-show-info.xlsx'.format(site))
    if 'sh_ver' in wb.get_sheet_names():
        ws = wb.get_sheet_by_name('sh_ver')
        wb.remove_sheet(ws)

    # Create sh_ver sheet and make active
    wb.create_sheet(title='sh_ver')
    sh_ver_sheet = wb.get_sheet_by_name('sh_ver')

    # For every file in Log Directory ending with .log
    for fn in os.listdir(os.getcwd()):

        # Get show ver header and show ver data
        if fn.endswith(".log"):
            sh_ver_header, sh_ver_data = log_file_to_show_ver(fn)

            # If row sheet length is 1, write header else write data
            if sh_ver_sheet.max_row == 1:
                sh_ver_sheet.append(sh_ver_header)

            # For each line in list write data excel row
            for line in sh_ver_data:
                sh_ver_sheet.append(line)

    # Save workbook
    sh_ver_sheet['I1'] = '=SUBTOTAL(3,H2:H{})'.format(sh_ver_sheet.max_row)
    sh_ver_sheet.auto_filter.ref = 'A1:H1'
    wb.save('{}-show-info.xlsx'.format(site))

    print 'SH VER TAB saved to:\t{}\{}-show-info.xlsx'.format(os.getcwd(), site)


def show_int_stat_to_excel(site, logfolder):

    # Change to Log File Directory
    os.chdir(logfolder)

    # Create site xlsx if not already created
    if not os.path.exists('{}-show-info.xlsx'.format(site)):
        wb = openpyxl.Workbook()
        wb.save('{}-show-info.xlsx')
        wb.close

    # Open workbook and delete sh_int_stat sheet if it exists
    wb = openpyxl.load_workbook(filename='{}-show-info.xlsx'.format(site))
    if 'sh_int_stat' in wb.get_sheet_names():
        ws = wb.get_sheet_by_name('sh_int_stat')
        wb.remove_sheet(ws)

    # Create sh_int_stat sheet and make active
    wb.create_sheet(title='sh_int_stat')
    sh_int_stat_sheet = wb.get_sheet_by_name('sh_int_stat')

    # For every file in Log Directory ending with .log
    for fn in os.listdir(os.getcwd()):

        # Get sh_ver header and data, sh_int_stat header and data
        if fn.endswith(".log"):
            sh_int_stat_header, sh_int_stat_data = log_file_to_show_int_stat(fn)
            sh_ver_header, sh_ver_data = log_file_to_show_ver(fn)

            # If row sheet length is 1, insert HOSTNAME to header and write header
            if sh_int_stat_sheet.max_row == 1:
                sh_int_stat_header.insert(0, sh_ver_header[2])
                sh_int_stat_sheet.append(sh_int_stat_header)

            # For each line in list insert <hostname> and write data to excel row
            for line in sh_int_stat_data:
                line.insert(0, sh_ver_data[0][2])
                sh_int_stat_sheet.append(line)

    # Save workbook
    sh_int_stat_sheet['I1'] = '=SUBTOTAL(3,H2:H{})'.format(sh_int_stat_sheet.max_row)
    sh_int_stat_sheet.auto_filter.ref = 'A1:H1'
    wb.save('{}-show-info.xlsx'.format(site))

    print 'SH INT STAT Tab saved to:\t{}\{}-show-info.xlsx'.format(os.getcwd(), site)


def show_cdp_nei_det_to_excel(site, logfolder):

    # Change to Log File Directory
    os.chdir(logfolder)

    # Create site xlsx if not already created
    if not os.path.exists('{}-show-info.xlsx'.format(site)):
        wb = openpyxl.Workbook()
        wb.save('{}-show-info.xlsx'.format(site))
        wb.close

    # Open workbook and delete sh_cdp_nei_det sheet if it exists
    wb = openpyxl.load_workbook(filename='{}-show-info.xlsx'.format(site))
    if 'sh_cdp_nei_det' in wb.get_sheet_names():
        ws = wb.get_sheet_by_name('sh_cdp_nei_det')
        wb.remove_sheet(ws)

    # Create sh_int_stat sheet and make active
    wb.create_sheet(title='sh_cdp_nei_det')
    sh_cdp_nei_det_sheet = wb.get_sheet_by_name('sh_cdp_nei_det')

    # For every file in Log Directory ending with .log
    for fn in os.listdir(os.getcwd()):

        # Get sh_cdp_nei_det header and data, sh_int_stat header and data
        if fn.endswith(".log"):
            sh_cdp_nei_det_header, sh_cdp_nei_det_data = log_file_to_show_cdp_nei_det(fn)
            sh_ver_header, sh_ver_data = log_file_to_show_ver(fn)

            # If row sheet length is 1, insert HOSTNAME to header and write header
            if sh_cdp_nei_det_sheet.max_row == 1:
                sh_cdp_nei_det_header.insert(0, sh_ver_header[2])
                sh_cdp_nei_det_sheet.append(sh_cdp_nei_det_header)

            # For each line in list insert <hostname> and write data to excel row
            for line in sh_cdp_nei_det_data:
                line.insert(0, sh_ver_data[0][2])
                sh_cdp_nei_det_sheet.append(line)

    # Save workbook
    # sh_cdp_nei_det_sheet['I1'] = '=SUBTOTAL(3,H2:H{})'.format(sh_cdp_nei_det_sheet.max_row)
    sh_cdp_nei_det_sheet.auto_filter.ref = 'A1:G1'
    wb.save('{}-show-info.xlsx'.format(site))

    print 'CDP NEI DET Tab saved to:\t{}\{}-show-info.xlsx'.format(os.getcwd(), site)


def log_folder_matrix_to_excel(site, logfolder):

    # Change to Log File Directory
    os.chdir(logfolder)

    # Create site xlsx if not already created
    if not os.path.exists('{}-show-info.xlsx'.format(site)):
        wb = openpyxl.Workbook()
        wb.save('{}-show-info.xlsx'.format(site))
        wb.close

    # Open workbook and delete sh_cdp_nei_det sheet if it exists
    wb = openpyxl.load_workbook(filename='{}-show-info.xlsx'.format(site))
    if 'conn_matrix' in wb.get_sheet_names():
        ws = wb.get_sheet_by_name('conn_matrix')
        wb.remove_sheet(ws)

    # Create conn_matix sheet and make active
    wb.create_sheet(title='conn_matrix')
    ws = wb.get_sheet_by_name('conn_matrix')

    for line in log_folder_to_matrix(logfolder):
        ws.append(line)
    ws.auto_filter.ref = 'A1:U1'
    wb.save('{}-show-info.xlsx'.format(site))

    print 'Matrix Tab saved to:\t{}\{}-show-info.xlsx'.format(os.getcwd(), site)


def main():
    # Timestamp the start of the run so that a total run time can be calculated at the end
    start_time = datetime.datetime.now()
    # ************************************************************************************
    show_version_to_excel(sys.argv[1], sys.argv[2])
    show_int_stat_to_excel(sys.argv[1], sys.argv[2])
    show_cdp_nei_det_to_excel(sys.argv[1], sys.argv[2])
    log_folder_matrix_to_excel(sys.argv[1], sys.argv[2])

    # **************************************************************************************
    # End - Calculate time of execution
    delta_time = datetime.datetime.now() - start_time
    textt = "Script Execution Time (s): " + str(delta_time.total_seconds())
    texttm = "Script Execution Time (m): " + str(delta_time.total_seconds() / 60)

    # Print script elapsed time
    print "-" * len(textt)
    print textt
    if delta_time.total_seconds() > 60:
        print texttm
    print "-" * len(textt)
    # ===================================================


if __name__ == "__main__":
    main()